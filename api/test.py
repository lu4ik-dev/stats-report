import openpyxl
from random import randint
from datetime import datetime

# Открыть шаблон Excel
template_path = "exp_template.xlsx"
wb = openpyxl.load_workbook(template_path)
ws = wb.active

# Записать "ГАПОУ 'БНК'" в ячейки A8-A29
for row in ws.iter_rows(min_row=8, max_row=29, min_col=1, max_col=1):
    for cell in row:
        cell.value = "ГАПОУ 'БНК'"

# Записать рандомные числа от 0 до 20 в ячейки D8-R29
for row in ws.iter_rows(min_row=8, max_row=29, min_col=4, max_col=18):
    for cell in row:
        cell.value = randint(0, 20)

# Сохранить шаблон с текущей датой в имени файла
date = datetime.now().strftime("%Y-%m-%d")
file_name = f"res_{date}.xlsx"
wb.save(file_name)

print(f"Файл успешно сохранен как {file_name}")
